from pdfminer.high_level import extract_text
from decimal import Decimal
import openpyxl
import re
from collections import defaultdict
import os
from fastapi.middleware.cors import CORSMiddleware
import uuid


def regex(text):
    pattern4 = r"^(?!.*\?).*\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern3 = r"^\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern2 = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+"
    lines = text.split("\n")
    matching_indices = []  # List to store indices of matching lines
    for lineIndex in range(len(lines)):
        if re.match(pattern4, lines[lineIndex]) or re.match(r"Totals\s*:\s*([\d.]+)" , lines[lineIndex]):  # If the pattern matches
            matching_indices.append(lineIndex)  # Add the index to the list
    return matching_indices  # Return the list of indices

def findLine(text , search):
    for lineIndex in range(len(text.split("\n"))):
        line = text.split("\n")[lineIndex]
        if search in line:
            return lineIndex

def normalize_key(key: str) -> str:
    # Convert to lowercase
    normalized_key = key.lower()

    # Remove all non-alphanumeric characters (except for numbers)
    normalized_key = re.sub(r'[^a-z0-9]', '', normalized_key)

    return normalized_key

def extract(pdf_file_path):
    titles = ["" for _ in range(5)]
    result = []
    with open(pdf_file_path, 'rb') as fh:
        for page_text in extract_text(fh).split("\f")[:-1]:
            sampleName = page_text.split("\n")[findLine(page_text , "Sample Name")].split(" ")
            sampleName = [x for x in sampleName if x]
            sampleName = normalize_key(sampleName[2])

            injections = []
            try:
                indecies = regex(page_text)
                print(sampleName)
                for i in range(len(indecies)):
                    areaLine = page_text.split("\n")[indecies[i]].split(" ")
                    areaLine = [x for x in areaLine if x]
                    areaLine2 = [x for x in areaLine if x.replace('.', '', 1).isdigit() and x.count('.') <= 1]
                    print(areaLine)
                    try:
                        Area = areaLine2[3]
                        match = re.search(r"\s+([A-Za-z\s]+)\s*$" , " ".join(areaLine))
                        if match: titles[i] = match.group(1)

                    except:
                        Area = areaLine[2]
                        titles[-1] = "Totals"

                    injections.append(Area)
            except Exception as e:
                print("skipped" , e)
                Area = -1

            for i in range(len(injections)):
                if len(result) < i + 1:
                    result.append(defaultdict(list))
                result[i][sampleName].append(injections[i])
    titles = [x for x in titles if len(x)]
    return [titles , result]

def push(extractedData):

    dicOG = {
        "st50": [6, 3],
        "st80": [6, 4],
        "st100": [6, 5],
        "st160": [6, 6],
        "st200": [6, 7],
        "t80": [18, 4],
        "t100": [18, 5],
        "t160": [18, 6],
        "f1": [32, 7],
        "f2": [32, 9],
        "sday": [58, 8],
        "sanalyst": [32, 4],
        "scolumn": [46, 9],
        "m2": [45, 5],
        "m1": [45, 3],
        "stability": [64, 3]
    }


    workbook = openpyxl.load_workbook('template3.xlsx')
    sheetOG = workbook.active

    extractedData[1] = extractedData[1][:len(extractedData[0])]

    for i in range(len(extractedData[1])):

        dic = {key: value[:] for key, value in dicOG.items()}

        if i == 0:
            sheet = sheetOG
            sheet.title = extractedData[0][i]
            sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData[0][i]}"
        else:
            sheet = workbook.copy_worksheet(sheetOG)
            sheet.title = extractedData[0][i]
            sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData[0][i]}"

        for key, values in extractedData[1][i].items():
            if key not in dic: continue
            for value in values:
                # print(key , value , dic[key][0] , dic[key][1])
                try:
                    sheet.cell(row=dic[key][0], column=dic[key][1]).value = Decimal(value)
                except:
                    print(f"Error at {key} {value} {dic[key][0]} {dic[key][1]}")
                dic[key][0] += 1

    workbook.save('result.xlsx')

push(extract("vald6.pdf"))
print("done")