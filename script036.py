from pdfminer.high_level import extract_text
from decimal import Decimal
import openpyxl
import re
from collections import defaultdict

def regex(text):
    pattern4 = r"^(?!.*\?).*\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern3 = r"^\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern2 = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+"
    lines = text.split("\n")
    matching_indices = []  # List to store indices of matching lines
    for lineIndex in range(len(lines)):
        if re.match(pattern4, lines[lineIndex]):  # If the pattern matches
            matching_indices.append(lineIndex)  # Add the index to the list
    return matching_indices  # Return the list of indices

def findLine(text , search):
    for lineIndex in range(len(text.split("\n"))):
        line = text.split("\n")[lineIndex]
        if search in line:
            return lineIndex

def extract(pdf_path):
    standard = []
    batches = defaultdict(list)
    with open(pdf_path, "rb") as fh:
        for page_text in extract_text(fh).split("\f"):
                sampleNameLineIndex = findLine(page_text , "Sample Name")
                if sampleNameLineIndex:
                    sampleNameLine = page_text.split("\n")[findLine(page_text , "Sample Name")]
                else: continue

                sampleName = re.search(r"Sample Name:\s*(.+)", sampleNameLine).group(1)

                lineIndex = regex(page_text)
                areaLine = page_text.split("\n")[lineIndex[0]].split(" ")
                areaLine = [x for x in areaLine if x.replace('.', '', 1).isdigit() and x.count('.') <= 1]
                Area = areaLine[3]
                if sampleName.startswith("st-"):
                    standard.append(Area)
                else:
                    batches[sampleName].append(Area)
    return {"standard" : standard ,"batches": batches}


def push(pdf_path , number):
    extractedData = extract(pdf_path)

    standard = extractedData["standard"]
    batches = extractedData["batches"]

    standardStart = [5 , 2]
    batchNameStart = [3 , 4]
    areaStart = [4 , 5]
    ogareaStartCol = 5

    if number == "3" :
        standardStart = [18 , 2]
        batchNameStart = [16 , 4]
        areaStart = [17 , 5]
    elif number == "6":
        standardStart = [31 , 2]
        batchNameStart = [29 , 4]
        areaStart = [30 , 5]



    workbook = openpyxl.load_workbook('template036.xlsx')
    sheetOG = workbook.active

    # print(sheetOG.cell(row=5 , column=5).value)

    for area in standard:
        try:
            sheetOG.cell(row=standardStart[0] , column=standardStart[1]).value = Decimal(area)
            print(f"inserted {area} at {standardStart}")
            standardStart[0] += 1
        except:
            # print(sheetOG.cell(row=7 , column=7).value)
            print("error at " + standardStart)
    for batchName , value in batches.items():
        try:
            sheetOG.cell(row=batchNameStart[0] , column=batchNameStart[1]).value = batchName
            print(f"inserted {batchName} at {batchNameStart}")
            batchNameStart[0] += 2
        except:
            print("error at " + batchNameStart)

        for area in value:
            try:
                sheetOG.cell(row=areaStart[0] , column=areaStart[1]).value = Decimal(area)
                print(f"inserted {area} at {areaStart}")
                areaStart[1] += 1
            except:
                print("error at " + areaStart)
        areaStart[1] = ogareaStartCol
        areaStart[0] += 2

    workbook.save("result036.xlsx")

push("6m.pdf" , "3")