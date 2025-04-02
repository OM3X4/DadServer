from pdfminer.high_level import extract_text
import re
from decimal import Decimal
import openpyxl
from collections import defaultdict

pattern = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+\s+\d+\.\d+(?:\s+\w+)?\s*$"
pattern2 = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+"

def regex(text):
    lines = text.split("\n")
    for lineIndex in range(len(lines)):
        if re.match(pattern2, lines[lineIndex]):
            return lineIndex
    return None


def findLine(text , search):
    for lineIndex in range(len(text.split("\n"))):
        line = text.split("\n")[lineIndex]
        if search in line:
            return lineIndex

def normalize_key(key: str) -> str:
    # Convert to lowercase
    normalized_key = key.lower()

    # Remove all non-alphanumeric characters (except for numbers)
    normalized_key = re.sub(r'[^a-z0-9]', '', normalized_key)

    return normalized_key

# row first
direction = {
    "st50": [5 , 2],
    "st80": [5 , 3],
    "st100": [5 , 4],
    "st160": [5 , 5],
    "st200": [5 , 6],
    "t80": [16 , 3],
    "t100": [16 , 4],
    "t160": [16 , 5],
    "scolumn": [16 , 5],
    "scolumn": [16 , 5],
}



def extract(pdf_file_path):
    extractedData = defaultdict(list)
    with open(pdf_file_path, 'rb') as fh:
        for page_text in extract_text(fh).split("\f")[:-1]:
            sampleName = page_text.split("\n")[findLine(page_text , "Sample Name")].split(" ")
            sampleName = [x for x in sampleName if x]
            sampleName = normalize_key(sampleName[2])
            # areaLine = page_text.split("\n")[findLine(page_text , "RetTime") + 3].split(" ")


            try:
                areaLine = page_text.split("\n")[regex(page_text)].split(" ")
                print(areaLine)
                areaLine = [x for x in areaLine if x]
                Area = areaLine[4]
                match = re.search(r"\s+([A-Za-z\s]+)\s*$" , " ".join(areaLine))
                if match: extractedData["title"] = f"CALCULATION OF VALIDATION OF  {match.group(1)}"
            except:
                print("Skipped")
                # areaLine = page_text.split("\n")[findLine(page_text , "RetTime") + 4].split(" ")
                # print("Skipped")
                # print(areaLine)
                Area = -1
            extractedData[sampleName].append(Area)
    return extractedData

# print(extract("vald4.pdf"))

def push(extractedData):

    dic = {
        "st50": [6, 3],
        "st80": [6, 4],
        "st100": [6, 5],
        "st160": [6, 6],
        "st200": [6, 7],
        "t80": [18, 4],
        "t100": [18, 5],
        "t160": [18, 6],
        "f1": [32, 7],
        "f2": [32, 9],
        "sday": [58, 8],
        "sanalyst": [32, 4],
        "scolumn": [46, 9],
        "m2": [45, 5],
        "m1": [45, 3],
        "stability": [64, 3]
    }


    workbook = openpyxl.load_workbook('template3.xlsx')
    sheet = workbook.active

    for key, values in extractedData.items():
        if key not in dic: continue
        for value in values:
            print(key , value , dic[key][0] , dic[key][1])
            try:
                sheet.cell(row=dic[key][0], column=dic[key][1]).value = Decimal(value)
            except:
                print(f"Error at {key} {value} {dic[key][0]} {dic[key][1]}")
            dic[key][0] += 1

    sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData['title']}"
    workbook.save('result.xlsx')


push(extract("vald5.pdf"))
# extract("vald2.pdf")
# findLine("vald2.pdf" , "Sample Name")

# with open("vald4.pdf", 'rb') as fh:
#         for page_text in extract_text(fh).split("\f")[:-1]:
#             # areaLine = page_text.split("\n")[findLine(page_text , "RetTime") + 3].split(" ")
#             areaLine = page_text.split("\n")[findLine(page_text , "RetTime") + 3].split(" ")
#             areaLine = [x for x in areaLine if x]
#             print(areaLine)

