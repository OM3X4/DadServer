from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse
from pdfminer.high_level import extract_text
from decimal import Decimal
import openpyxl
import re
from collections import defaultdict
import os
from fastapi.middleware.cors import CORSMiddleware
import uuid
from docsApp import pushToWord

from app036 import push as push036


pattern = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+\s+\d+\.\d+(?:\s+\w+)?\s*$"
pattern2 = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+"

def regex(text):
    pattern4 = r"^(?!.*\?).*\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern3 = r"^\s*\d+\s+\d+\.\d+\s+[\w\s]+\s+\d+\.\d+\s+\d+\.\d+"
    pattern2 = r"^\s*\d+\s+\d+\.\d+\s+\w+\s+\d+\.\d+\s+\d+\.\d+"
    lines = text.split("\n")
    matching_indices = []  # List to store indices of matching lines
    for lineIndex in range(len(lines)):
        if re.match(pattern4, lines[lineIndex]) or re.match(r"Totals\s*:\s*([\d.]+)" , lines[lineIndex]):  # If the pattern matches
            matching_indices.append(lineIndex)  # Add the index to the list
    return matching_indices  # Return the list of indices

def findLine(text , search):
    for lineIndex in range(len(text.split("\n"))):
        line = text.split("\n")[lineIndex]
        if search in line:
            return lineIndex

def normalize_key(key: str) -> str:
    # Convert to lowercase
    normalized_key = key.lower()

    # Remove all non-alphanumeric characters (except for numbers)
    normalized_key = re.sub(r'[^a-z0-9]', '', normalized_key)

    return normalized_key

def extract(pdf_file_path):
    titles = ["" for _ in range(5)]
    result = []
    with open(pdf_file_path, 'rb') as fh:
        for page_text in extract_text(fh).split("\f")[:-1]:
            sampleName = page_text.split("\n")[findLine(page_text , "Sample Name")].split(" ")
            sampleName = [x for x in sampleName if x]
            sampleName = normalize_key(sampleName[2])

            injections = []
            try:
                indecies = regex(page_text)
                print(sampleName)
                for i in range(len(indecies)):
                    areaLine = page_text.split("\n")[indecies[i]].split(" ")
                    areaLine = [x for x in areaLine if x]
                    areaLine2 = [x for x in areaLine if x.replace('.', '', 1).isdigit() and x.count('.') <= 1]
                    print(areaLine)
                    try:
                        Area = areaLine2[3]
                        match = re.search(r"\s+([A-Za-z\s]+)\s*$" , " ".join(areaLine))
                        if match: titles[i] = match.group(1)

                    except:
                        Area = areaLine[2]
                        titles[-1] = "Totals"

                    injections.append(Area)
            except Exception as e:
                print("skipped" , e)
                Area = -1

            for i in range(len(injections)):
                if len(result) < i + 1:
                    result.append(defaultdict(list))
                result[i][sampleName].append(injections[i])
    titles = [x for x in titles if len(x)]
    return [titles , result]

def push(extractedData):

    dicOG = {
        "st50": [6, 3],
        "st80": [6, 4],
        "st100": [6, 5],
        "st160": [6, 6],
        "st200": [6, 7],
        "t80": [18, 4],
        "t100": [18, 5],
        "t160": [18, 6],
        "f1": [32, 7],
        "f2": [32, 9],
        "sday": [58, 8],
        "sanalyst": [32, 4],
        "scolumn": [46, 9],
        "m2": [45, 5],
        "m1": [45, 3],
        "stability": [64, 3]
    }


    workbook = openpyxl.load_workbook('template3.xlsx')
    sheetOG = workbook.active

    extractedData[1] = extractedData[1][:len(extractedData[0])]

    for i in range(len(extractedData[1])):

        dic = {key: value[:] for key, value in dicOG.items()}

        if i == 0:
            sheet = sheetOG
            sheet.title = extractedData[0][i]
            sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData[0][i]}"
        else:
            sheet = workbook.copy_worksheet(sheetOG)
            sheet.title = extractedData[0][i]
            sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData[0][i]}"

        for key, values in extractedData[1][i].items():
            if key not in dic: continue
            for value in values:
                # print(key , value , dic[key][0] , dic[key][1])
                try:
                    sheet.cell(row=dic[key][0], column=dic[key][1]).value = Decimal(value)
                except:
                    print(f"Error at {key} {value} {dic[key][0]} {dic[key][1]}")
                dic[key][0] += 1

    workbook.save('result.xlsx')
    return 'result.xlsx'


def regexOG(text):
    lines = text.split("\n")
    for lineIndex in range(len(lines)):
        if re.match(pattern2, lines[lineIndex]):
            return lineIndex
    return None


# Your existing script (with slight modifications if necessary)
def findLineOG(text , search):
    for lineIndex in range(len(text.split("\n"))):
        line = text.split("\n")[lineIndex]
        if search in line:
            return lineIndex


def normalize_keyOG(key: str) -> str:
    normalized_key = key.lower()
    normalized_key = re.sub(r'[^a-z0-9]', '', normalized_key)
    return normalized_key


def extractOG(pdf_file_path):
    extractedData = defaultdict(list)
    with open(pdf_file_path, 'rb') as fh:
        for page_text in extract_text(fh).split("\f")[:-1]:
            sampleName = page_text.split("\n")[findLineOG(page_text , "Sample Name")].split(" ")
            sampleName = [x for x in sampleName if x]
            sampleName = normalize_keyOG(sampleName[2])
            try:
                areaLine = page_text.split("\n")[regexOG(page_text)].split(" ")
                areaLine = [x for x in areaLine if x]
                Area = areaLine[4]
                match = re.search(r"\s+([A-Za-z\s]+)\s*$" , " ".join(areaLine))
                if match: extractedData["title"] = f"CALCULATION OF VALIDATION OF  {match.group(1)}"
            except:
                print("Skipped")
                Area = -1
            extractedData[sampleName].append(Area)
    return extractedData


def pushOG(extractedData, excel_template):
    dic = {
        "st50": [6, 3],
        "st80": [6, 4],
        "st100": [6, 5],
        "st160": [6, 6],
        "st200": [6, 7],
        "std50": [6, 3],
        "std80": [6, 4],
        "std100": [6, 5],
        "std160": [6, 6],
        "std200": [6, 7],
        "t80": [18, 4],
        "t100": [18, 5],
        "t160": [18, 6],
        "f1": [32, 7],
        "f2": [32, 9],
        "sday": [58, 8],
        "sanalyst": [32, 4],
        "scolumn": [46, 9],
        "m2": [45, 5],
        "m1": [45, 3],
        "stability": [64, 3]
    }

    workbook = openpyxl.load_workbook(excel_template)
    sheet = workbook.active

    for key, values in extractedData.items():
        if key not in dic:
            print(f"Key {key} not found in dictionary, skipping.")
            continue

        for value in values:
            print(f"Writing {key} value {value} to row {dic[key][0]} and column {dic[key][1]}")
            try:
                sheet.cell(row=dic[key][0], column=dic[key][1]).value = Decimal(value)
                dic[key][0] += 1
            except Exception as e:
                print(f"Error writing to Excel for {key}: {e}")

    result_filename = 'result.xlsx'
    sheet.cell(1 , 4).value = f"CALCULATION OF VALIDATION OF {extractedData['title']}"
    workbook.save(result_filename)
    return result_filename

app = FastAPI()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # You can replace '*' with specific domains if needed
    allow_credentials=True,
    allow_methods=["*"],  # Allow all HTTP methods (GET, POST, etc.)
    allow_headers=["*"],  # Allow all headers
)


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    # Temporary file to store the uploaded PDF
    pdf_path = f"temp_{uuid.uuid4().hex}.pdf"

    with open(pdf_path, "wb") as f:
        f.write(await file.read())

    try:
        # Extract data from the uploaded PDF
        extracted_data = extract(pdf_path)

        # Process the extracted data and push it into the Excel template
        result_file_name = push(extracted_data)

        # Return the result Excel file as a response
        return FileResponse(result_file_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="result.xlsx")

    finally:
        # Clean up the temporary PDF file
        os.remove(pdf_path)


@app.post("/uploadog")
async def upload_file(file: UploadFile = File(...)):
    # Temporary file to store the uploaded PDF
    pdf_path = f"temp_{uuid.uuid4().hex}.pdf"

    with open(pdf_path, "wb") as f:
        f.write(await file.read())

    try:
        # Extract data from the uploaded PDF
        extracted_data = extractOG(pdf_path)

        # Process the extracted data and push it into the Excel template
        result_file_name = pushOG(extracted_data , "template3.xlsx")

        # Return the result Excel file as a response
        return FileResponse(result_file_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="result.xlsx")

    finally:
        # Clean up the temporary PDF file
        os.remove(pdf_path)


@app.post("/036")
async def upload_file_036(file: UploadFile = File(...) , number: str = "0"):
    # Temporary file to store the uploaded PDF
    pdf_path = f"temp_{uuid.uuid4().hex}.pdf"

    with open(pdf_path, "wb") as f:
        f.write(await file.read())

    try:

        # Process the extracted data and push it into the Excel template
        result_file_name = push036(pdf_path , number)

        # Return the result Excel file as a response
        return FileResponse(result_file_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="result.xlsx")

    finally:
        # Clean up the temporary PDF file
        os.remove(pdf_path)


@app.post("/word")
async def word(file: UploadFile = File(...)):
    # Temporary file to store the uploaded PDF
    word_path = f"temp_{uuid.uuid4().hex}.docx"

    with open(word_path, "wb") as f:
        f.write(await file.read())

    try:

        # Process the extracted data and push it into the Excel template
        result_file_name = pushToWord(word_path)

        # Return the result Excel file as a response
        return FileResponse(result_file_name, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", filename="resultWord.docx")

    finally:
        # Clean up the temporary PDF file
        os.remove(word_path)

