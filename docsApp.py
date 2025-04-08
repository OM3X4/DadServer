from docx import Document
import openpyxl

# workbook = openpyxl.load_workbook("example.xlsx")
# sheet = workbook.active
doc = Document("template2.docx")
workbook = openpyxl.load_workbook("result(36).xlsx" , data_only=True)
sheet = workbook.active

tables = doc.tables

# print(table)



def pushToWord(FilePath):

    # first table
    startPositionWord = [1 , 1]
    startPositionExcel = [6 , 5]
    table = tables[0]
    for i in range(9):

        row = table.rows[startPositionWord[0]]

        row.cells[startPositionWord[1]].text = str(sheet.cell(startPositionExcel[0] , startPositionExcel[1]).value)
        startPositionWord[0] += 1
        startPositionExcel[0] += 1


    # #second Table
    word = [1 , 1]
    excel = [8 , 10]
    table = tables[1]
    for i in range(5):

        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [8 , 11]
    for i in range(8):

        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1


    #third Table
    table = tables[2]
    word = [1 , 1]
    excel = [18 , 5]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1


    ### Fourth Table
    table = tables[3]
    word = [1 , 0]
    excel = [21 , 9]
    for i in range(3):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value) + "mg/ml"
        word[0] += 3
        excel[0] += 3

    word = [1 , 1]
    excel = [21 , 10]
    for i in range(3):
        for i in range(9):
            row = table.rows[word[0]]
            row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
            word[0] += 1
            excel[0] += 1
        word[0] = 1
        excel[0] = 21
        word[1] += 1
        excel[1] += 1


    ## Table Five
    table = tables[4]
    excel = [6 , 5]
    word = [1 , 1]
    for i in range(7):
        try:
            row = table.rows[word[0]]
            row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
            word[0] += 1
            excel[0] += 1
        except Exception as e:
            print(e , excel , word)



    excel = [18 , 5]
    word = [1 , 2]
    for i in range(7):
        try:
            row = table.rows[word[0]]
            row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
            word[0] += 1
            excel[0] += 1
        except Exception as e:
            print(e , excel , word)


    ### Table six
    table = tables[5]
    excel = [58 , 7]
    word = [1 , 1]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [58 , 8]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    ### Table Seven
    table = tables[6]
    excel = [32 , 3]
    word = [1 , 1]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [32 , 4]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1




    ### Table Eight
    table = tables[7]
    excel = [46 , 8]
    word = [1 , 1]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [46 , 9]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1





    ### Table Nine
    table = tables[8]
    excel = [45 , 3]
    word = [1 , 1]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [45 , 4]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 3]
    excel = [45 , 5]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1



    ### Table Ten
    table = tables[9]
    excel = [32 , 7]
    word = [1 , 1]
    for i in range(9):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 2]
    excel = [32 , 8]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1

    word = [1 , 3]
    excel = [32 , 9]
    for i in range(6):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1


    ### Table Eleven
    table = tables[10]
    excel = [58 , 3]
    word = [1 , 1]
    for i in range(15):
        row = table.rows[word[0]]
        row.cells[word[1]].text = str(sheet.cell(excel[0] , excel[1]).value)
        word[0] += 1
        excel[0] += 1



    doc.save("resultWord.docx")
    return "resultWord.docx"